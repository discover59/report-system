import os

from openpyxl import Workbook


def get_file_content(filepath):
    line_data = []
    with open(filepath) as f:
        for line in f:
            line_data.append(line)

    return line_data


def get_non_compliant_list(file_path):
    file_input = get_file_content(file_path)
    arrays = list()
    pools = dict()
    tmp = dict()
    compliant = True
    for item in file_input:
        if '=======' in item:
            if pools != {} and not compliant:
                arrays.append(pools)
            compliant = True
            pools = dict(name=item.strip('=\n')[5:].strip(), list=[])
        elif 'Relocation Type' in item:
            tmp['Relocation Type'] = item.split('Relocation Type')[1].strip(': \n')
            if 'N/A' in tmp['Relocation Type']:
                compliant = False
            pools['list'].append(tmp)
            tmp = {}
        elif 'Storage Pool Name' in item:
            tmp['Storage Pool Name'] = item.split('Storage Pool Name')[1].strip(': \n')
        elif 'Storage Pool ID' in item:
            tmp['Storage Pool ID'] = item.split('Storage Pool ID')[1].strip(': \n')
    return arrays


if __name__ == "__main__":
    parsed_data = get_non_compliant_list(os.path.dirname(__file__) + '/test.txt')
    wb = Workbook()
    target_file = 'output.xlsx'
    ws = wb.active
    ws.title = 'output'
    ws.cell(column=1, row=1, value='Total Number of IP')
    ws.cell(column=2, row=1, value=len(parsed_data))
    ws.cell(column=1, row=3, value='Arrays that are not compliant w.r.t at least one pool')
    ws.cell(column=1, row=5, value='Array Name')
    ws.cell(column=2, row=5, value='IP')
    ws.cell(column=3, row=5, value='Storage Pool Name')
    ws.cell(column=4, row=5, value='Compliant Yes')
    ws.cell(column=5, row=5, value='Compliant No')
    cur_row = 6
    for item in parsed_data:
        ws.cell(column=2, row=cur_row, value=item['name'])
        for child in item['list']:
            ws.cell(column=3, row=cur_row, value=child['Storage Pool Name'])
            if 'N/A' in child['Relocation Type']:
                ws.cell(column=5, row=cur_row, value='NO')
            else:
                ws.cell(column=4, row=cur_row, value=child['Relocation Type'])
            cur_row += 1
        cur_row += 2

    wb.save(target_file)
