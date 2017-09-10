# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import datetime
import os.path

from io import BytesIO

from django.http.response import JsonResponse
from django.shortcuts import render
from django.views.generic import TemplateView, View

from openpyxl import load_workbook, utils as pyxl_utils

from .models import ExcelFile
from .utils import run_script


class DashboardPage(TemplateView):
    template_name = 'main.html'

    def dispatch(self, *args, **kwargs):
        return super(DashboardPage, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(DashboardPage, self).get_context_data(**kwargs)
        # if self.request.GET.get('date'):
        #     context['date'] = self.request.GET.get('date')
        # else:
        #     context['date'] = datetime.date.today
        # context['property'] = Property.objects.all().values()
        return context

    def get(self, request, *args, **kwargs):
        return render(request, 'main.html')


class ParseFileOld(View):
    """
    View for uploading excel data with date and getting excel data with date
    """
    def get(self, request):
        selected_date = datetime.datetime.strptime(request.GET.get('date'), '%a %b %d %Y')
        selected = ExcelFile.objects.filter(date=selected_date).first()
        if not selected:
            return JsonResponse(dict(state=False))
        file_path = selected.excel_file.path
        if file_path:
            wb = load_workbook(filename=file_path)
            # wb.get_sheet_names()
            response = {}
            sheet_names = wb.get_sheet_names()
            for item in sheet_names:
                if item == 'Introduction':
                    continue
                ws = wb.get_sheet_by_name(item)
                sheet_data = {}
                for m_range in ws.merged_cell_ranges:
                    merged_cells = list(pyxl_utils.rows_from_range(m_range))
                    table_title, table_data, array_data = get_table_data(ws, merged_cells)
                    sheet_data[table_title] = array_data
                response[item] = sheet_data

            return JsonResponse(dict(state=True, res=response))

        return JsonResponse(dict(state=False))

    def post(self, request, *args, **kwargs):
        uploaded = request.FILES.get('file', None)
        if not uploaded:
            return JsonResponse(dict(status=False))
        wb = load_workbook(filename=BytesIO(uploaded.read()))
        # wb.get_sheet_names()
        ws = wb.get_sheet_by_name('vCenter')
        values = []
        response = {}
        for m_range in ws.merged_cell_ranges:
            merged_cells = list(pyxl_utils.rows_from_range(m_range))
            table_title, table_data, array_data = get_table_data(ws, merged_cells)
            response[table_title] = array_data

        return JsonResponse(response)


class UploadFile(View):
    template_name = 'file_upload.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        uploaded = request.FILES.get('file', None)
        file_date = request.POST.get('date', '')
        try:
            file_obj, created = ExcelFile.objects.update_or_create(
                date=file_date,
                defaults={
                    'excel_file': uploaded
                }
            )
        except Exception as e:
            return JsonResponse(dict(status=False))

        if not created:
            return JsonResponse(dict(status=False))

        return JsonResponse(dict(status=True))


class ExcelView(View):
    """
    View for uploading excel data with date and getting excel data with date
    """
    def get(self, request):
        selected_date = datetime.datetime.fromtimestamp(float(request.GET.get('date')) / 1000)
        file_path = os.path.dirname(__file__) + '/../files/' + selected_date.strftime('%Y-%m-%d') + '.xlsx'
        try:
            wb = load_workbook(filename=file_path)
            # wb.get_sheet_names()
            response = {}
            sheet_names = wb.get_sheet_names()
            for item in sheet_names:
                if item == 'Introduction':
                    continue
                ws = wb.get_sheet_by_name(item)
                sheet_data = {}
                for m_range in ws.merged_cell_ranges:
                    merged_cells = list(pyxl_utils.rows_from_range(m_range))
                    table_title, table_data, array_data = get_table_data(ws, merged_cells)
                    sheet_data[table_title] = array_data
                response[item] = sheet_data

            return JsonResponse(dict(state=True, res=response))
        except Exception as e:
            return JsonResponse(dict(state=False))


class ScriptView(View):
    """
    View for getting script result
    """
    def post(self, request):
        return JsonResponse(dict(res=run_script()))


def get_table_data(worksheet, col_range):
    """
    returns sub titles and values corresponding to sub titles of selected table
    :param worksheet: selected sheet of excel file
    :param col_range: list of tuple of cell strings
    :param row_number:
    :return:
    """
    table_title = worksheet.cell(col_range[0][0]).value
    table_data = {}

    x, y = pyxl_utils.coordinate_from_string(col_range[0][0])
    first_col = pyxl_utils.column_index_from_string(x)
    sub_title_list = []
    for i in range(0, len(col_range[0])):
        sub_title = worksheet[pyxl_utils.get_column_letter(first_col + i) + str(y + 1)].value
        sub_title_list.append(sub_title)
        table_data[sub_title] = []
    table_data_array = [sub_title_list]
    b = True
    start_row = y + 2
    while b:
        b = False
        row_array = []
        for i in range(0, len(col_range[0])):
            cell_value = worksheet[pyxl_utils.get_column_letter(first_col + i) + str(start_row)].value
            if cell_value:
                b = True
            table_data[sub_title_list[i]].append(cell_value)
            row_array.append(cell_value)
        table_data_array.append(row_array)
        start_row += 1

    return table_title, table_data, table_data_array
